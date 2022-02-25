
from re import S
from attr import s
from click import option
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import openpyxl
from bs4 import BeautifulSoup
import time
import pyautogui
import os
from datetime import date

today = str(date.today())

if not(os.path.isdir(today)):
   os.makedirs(os.path.join(today))

wb = openpyxl.Workbook()
ws = wb.active
ws['a1'] = '제목'
ws['b1'] = '컨텐츠'
ws['e1'] = '키워드'

myorder = pyautogui.prompt("키워드 :")
ws['f1'] = myorder 

option = webdriver.ChromeOptions() 
option.add_argument('headless')

dv = webdriver.Chrome(options=option) #크롬 안보이게 숨기기
#dv.get("https://search.naver.com/search.naver?query=%s&nso=&where=view&sm=tab_viw.all&mode=normal" % myorder) 네이버 view 페이지
dv.get("https://search.naver.com/search.naver?where=blog&sm=tab_viw.blog&query=%s&nso=" % myorder) #블로그전용

# 검색한 화면에서 나온 1페이지 제목 뽑기
contents = dv.find_elements_by_css_selector('.api_txt_lines.total_tit')

#제목 크롤링
row = 1 
for content in contents:
    ws[f'a{row + 1}'] = content.text
    dv.implicitly_wait(10)
    row = row + 1
    #print(content.text)
    wb.save('컨텐츠.xlsx')

#클릭할 페이지 크롤링
i = 1 
for i in range(1,31):
    #elem = dv.find_element_by_xpath("//*[@id='main_pack']/section/div/div[2]/panel-list/div[1]/more-contents/div/ul/li[%d]/div/div/a" % i) #네이버 view
    elem = dv.find_element_by_xpath(f'//*[@id="sp_blog_{i}"]/div/div/a')                            
    elem.click()
    
    
# 크롬 브라우저 새 창활성화 시키기
    
    dv.switch_to.window(dv.window_handles[1]) #새 창 활성화
    dv.implicitly_wait(10)
    try:
        swframe = dv.find_element_by_id("mainFrame") #iframe 태그 엘리먼트 찾기
        dv.switch_to.frame(swframe) #프레임 이동\
        
    except: #블로그 이외의 포스팅은 제외
        dv.close()    
        dv.switch_to.window(dv.window_handles[0])
        i = i + 1
        elem = dv.find_element_by_xpath("//*[@id='main_pack']/section/div/div[2]/panel-list/div[1]/more-contents/div/ul/li[%d]/div/div/a" % i)
        elem.click()
        dv.switch_to.window(dv.window_handles[1])
    
#제목뽑기 고민 중인공간    

    elist = dv.find_elements_by_class_name('se-main-container')
    dv.implicitly_wait(10)  
    for e in elist:    
        ws[f'b{i + 1}'] = e.text
        wb.save('컨텐츠.xlsx')
        #print(e.text)
        dv.implicitly_wait(10)

    dv.close() #새 창 끄기
    dv.switch_to.window(dv.window_handles[0]) #원래 메인창으로 활성화 복귀

    
dv.close()
dv.quit()
pyautogui.alert("완료")