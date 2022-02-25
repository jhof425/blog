
from argparse import Action
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
import time
import pyautogui
import pyperclip
from openpyxl import load_workbook

lb = load_workbook('컨텐츠.xlsx', data_only=True)
ws = lb.active


id = 'ancozo89'
pw = '8aowlr9'

dv = webdriver.Chrome()
dv.get("https://nid.naver.com/nidlogin.login")
action = ActionChains(dv)

dv.implicitly_wait(10)

elem = dv.find_element_by_id('id').click()
pyperclip.copy(id)
pyautogui.hotkey('ctrl', 'v')
dv.implicitly_wait(10)

elem = dv.find_element_by_id('pw').click()
pyperclip.copy(pw)
pyautogui.hotkey('ctrl', 'v')
dv.implicitly_wait(10)

elem = dv.find_element_by_xpath('//*[@id="log.login"]').click()
dv.implicitly_wait(10)

elem = dv.find_element_by_xpath('//*[@id="NM_FAVORITE"]/div[1]/ul[1]/li[3]/a').click()
dv.implicitly_wait(10)

elem = dv.find_element_by_xpath('//*[@id="container"]/div/aside/div/div[1]/nav/a[2]').click()
time.sleep(3)

dv.switch_to.window(dv.window_handles[1]) #새 창 활성화
swframe = dv.find_element_by_id("mainFrame") #iframe 태그 엘리먼트 찾기
dv.switch_to.frame(swframe) #프레임 이동\
dv.implicitly_wait(10)

dv.find_element_by_xpath('//span[contains(text(),"제목")]').click()
action.send_keys(ws['a2'].value).perform() #openxl 제목 옮겨 적기

dv.find_element_by_xpath('//span[contains(text(),"본문에")]').click()
time.sleep(1)
action = ActionChains(dv)
time.sleep(1)
action.send_keys(ws['b2'].value).perform()
time.sleep(1)