from types import NoneType
from openpyxl import load_workbook

wb = load_workbook('컨텐츠.xlsx')
ws = wb.active

get_sub = ws['a1'].value
get_name = ws['f1'].value

f = open(f'{get_name}' + f"{get_sub}.txt", 'w', encoding='utf-8')
for i in range(2, 32):
    data = ws[f'a{i}'].value + "\n" 
    #data = "%d번째 줄입니다.\n" % i
    f.write(data)
f.close()

f = open(f"{get_name}.txt", 'w', encoding='utf-8')
for i in range(2, 32):
    try:
        data = ws[f'b{i}'].value + "\n"     
        #data = "%d번째 줄입니다.\n" % i
        f.write(data)
    except TypeError:
        continue

f.close()
    
